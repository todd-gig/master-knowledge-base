#!/usr/bin/env python3
"""
xlsx_to_knowledge.py
────────────────────
Converts project-associated Excel workbooks into structured Markdown knowledge
files consumable by AI intelligence systems (Claude, GPT, RAG pipelines).

Usage:
    python scripts/xlsx_to_knowledge.py [--output-dir OUTPUT_DIR]

Output:
    One .md file per workbook, written to OUTPUT_DIR (default: ./knowledge-extracts/)

Supported workbooks (auto-detected by filename pattern):
    Sales_Operating_System.xlsx        → sales-operating-system project
    Brand_Voice_Data_Dictionary.xlsx   → master-knowledge-base / brand
    gigaton_playa_roisummary.xlsx      → gigaton / carmen-beach DAG model
    CarmenBeach_Pricing_Calculator.xlsx → carmen-beach service catalog
    CarmenBeach_Costing_Calculator_Governed.xlsx → carmen-beach agent costing
    CarmenBeach_Affiliate_Growth_Calculator.xlsx → carmen-beach affiliate model
    Occupancy_Marketing_ROI_Calculator.xlsx      → carmen-beach property ROI
    LiqueFex_Scenario_Forecasting_Calculator.xlsx → liquifex portfolio
    LiquiFex_Pilot1_Calculator_Restructured.xlsx  → liquifex pilot
    company_valuation_model.xlsx        → gigaton company valuation
    claude_tool_education_roi_calculator.xlsx → claude training ROI
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


# ─────────────────────────────────────────────────────────────────────────────
# Workbook registry — maps filename patterns to metadata
# ─────────────────────────────────────────────────────────────────────────────

WORKBOOK_REGISTRY: List[Dict[str, Any]] = [
    {
        "pattern": "Sales_Operating_System",
        "project": "sales-operating-system",
        "slug": "sales-os-catalog",
        "description": "214-item product/service master catalog with upsell, cross-sell, bundle, and scoring logic",
        "tags": ["sales", "catalog", "upsell", "cross-sell", "scoring"],
        "key_sheets": ["Master_Catalog", "Upsell_Matrix", "Cross_Sell_Matrix", "Bundles", "Client_Needs_Mapping", "Scoring_Model"],
        "max_rows": 250,
    },
    {
        "pattern": "Brand_Voice_Data_Dictionary",
        "project": "master-knowledge-base",
        "slug": "brand-voice-data-dictionary",
        "description": "200+ brand intelligence questions covering company identity, brand strategy, products/services, audiences, and competitive landscape",
        "tags": ["brand", "voice", "identity", "strategy", "audience", "competitive"],
        "key_sheets": ["Data_Dictionary", "Reference"],
        "max_rows": 300,
    },
    {
        "pattern": "gigaton_playa_roisummary",
        "project": "gigaton-engine",
        "slug": "gigaton-playa-dag-model",
        "description": "Causal DAG model for Playa del Carmen property ROI: exogenous market inputs, operator-controlled treatments, and outcome coefficients",
        "tags": ["dag", "causal", "playa-del-carmen", "roi", "occupancy", "revenue", "gigaton"],
        "key_sheets": ["DAG_Nodes", "Conv_Coeffs", "Occ_Coeffs", "Rev_Coeffs", "Scenarios", "Channel_Scenario", "Owner_Summary"],
        "max_rows": 100,
    },
    {
        "pattern": "CarmenBeach_Pricing_Calculator",
        "project": "Carmen-Beach-Properties",
        "slug": "carmen-beach-pricing",
        "description": "Service pricing catalog for Carmen Beach: media capture, listing optimization, Gigaton tech stack integration, property management packages",
        "tags": ["carmen-beach", "pricing", "service-catalog", "gigaton-integration"],
        "key_sheets": ["ServiceCatalog", "Packages", "Pricing", "AutomationLogic"],
        "max_rows": 100,
    },
    {
        "pattern": "CarmenBeach_Costing_Calculator_Governed",
        "project": "Carmen-Beach-Properties",
        "slug": "carmen-beach-agent-costing",
        "description": "Governed compensation model for Carmen Beach AI agent roles: Photographer, Listing Agent, Leasing Agent, Property Management Agent, Tenant Management",
        "tags": ["carmen-beach", "costing", "agent-compensation", "governance"],
        "key_sheets": ["Config", "Governance", "Causal_Matrix", "Summary_Role", "Summary_Outcome"],
        "max_rows": 100,
    },
    {
        "pattern": "CarmenBeach_Affiliate_Growth_Calculator",
        "project": "Carmen-Beach-Properties",
        "slug": "carmen-beach-affiliate-model",
        "description": "Affiliate growth and revenue projection model for Carmen Beach STR/LTR property acquisition",
        "tags": ["carmen-beach", "affiliate", "revenue-model", "str", "ltr"],
        "key_sheets": ["Inputs", "Scenarios"],
        "max_rows": 80,
    },
    {
        "pattern": "Occupancy_Marketing_ROI_Calculator",
        "project": "Carmen-Beach-Properties",
        "slug": "carmen-beach-occupancy-roi",
        "description": "Occupancy and marketing ROI model for Playa del Carmen properties (ADR $180, baseline 60% occupancy)",
        "tags": ["carmen-beach", "occupancy", "marketing-roi", "adr"],
        "key_sheets": ["Inputs", "Calc", "Scenarios"],
        "max_rows": 60,
    },
    {
        "pattern": "LiqueFex_Scenario_Forecasting",
        "project": "liquifex",
        "slug": "liquifex-scenario-forecasting",
        "description": "Life insurance policy portfolio scenario forecasting: subscription fee rates, preferred return rates, debt costs across base/downside/upside scenarios",
        "tags": ["liquifex", "portfolio", "scenario", "life-insurance", "forecasting"],
        "key_sheets": ["Scenarios", "Inputs", "Portfolio", "Outputs"],
        "max_rows": 80,
    },
    {
        "pattern": "LiquiFex_Pilot1_Calculator",
        "project": "liquifex",
        "slug": "liquifex-pilot1-roi",
        "description": "LiquiFex Pilot 1 ROI calculator: $500M AUM portfolio across 4, 6, and 12-month liquidity scenarios",
        "tags": ["liquifex", "pilot", "roi", "aum", "liquidity"],
        "key_sheets": ["Inputs", "4 Month", "6 Month", "12 Month"],
        "max_rows": 80,
    },
    {
        "pattern": "company_valuation_model",
        "project": "gigaton",
        "slug": "gigaton-company-valuation",
        "description": "Gigaton company valuation model: Playa del Carmen TAM ~$985M, 1.5% market share target, 30% operating margin",
        "tags": ["gigaton", "valuation", "tam", "playa-del-carmen", "market-share"],
        "key_sheets": ["Company Valuation", "Playa del Carmen Scenario"],
        "max_rows": 60,
    },
    {
        "pattern": "claude_tool_education_roi",
        "project": "master-knowledge-base",
        "slug": "claude-tool-education-roi",
        "description": "ROI calculator for Claude/AI tool education programs: workforce productivity, adoption rates, and net benefit computation",
        "tags": ["claude", "ai-training", "roi", "workforce", "education"],
        "key_sheets": ["ROI Calculator", "README"],
        "max_rows": 60,
    },
]


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return f"{val:.4g}"
    return str(val)


def _row_to_md_cells(row: Tuple, max_cols: int = 8) -> str:
    cells = [_cell_str(c) for c in row[:max_cols]]
    return "| " + " | ".join(cells) + " |"


def _separator(ncols: int) -> str:
    return "| " + " | ".join(["---"] * ncols) + " |"


def _sheet_to_markdown(ws, max_rows: int = 100, max_cols: int = 8) -> str:
    rows = []
    header_written = False
    ncols = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            rows.append(f"\n_... (truncated at {max_rows} rows)_")
            break
        if not any(c for c in row):
            if rows:
                rows.append("")
            continue
        visible = row[:max_cols]
        ncols = max(ncols, sum(1 for c in visible if c is not None))

        if not header_written:
            rows.append(_row_to_md_cells(visible, max_cols))
            rows.append(_separator(len(visible)))
            header_written = True
        else:
            rows.append(_row_to_md_cells(visible, max_cols))

    return "\n".join(rows)


def _find_workbook(pattern: str, search_paths: List[Path]) -> Optional[Path]:
    for base in search_paths:
        for f in base.glob("**/*.xlsx"):
            if pattern.lower() in f.stem.lower():
                return f
    return None


def _match_registry(path: Path) -> Optional[Dict[str, Any]]:
    stem = path.stem.lower()
    for entry in WORKBOOK_REGISTRY:
        if entry["pattern"].lower() in stem:
            return entry
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Core extraction
# ─────────────────────────────────────────────────────────────────────────────

def extract_workbook(path: Path, meta: Dict[str, Any]) -> str:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    tags_str = "\n".join(f"  - {t}" for t in meta["tags"])
    header = f"""---
title: {meta['slug']}
project: {meta['project']}
source_file: {path.name}
description: {meta['description']}
extracted: {date.today().isoformat()}
tags:
{tags_str}
type: xlsx-knowledge-extract
---

# {meta['description']}

**Source**: `{path.name}`
**Project**: `{meta['project']}`
**Extracted**: {date.today().isoformat()}

"""

    sections: List[str] = [header]
    available = wb.sheetnames
    target_sheets = [s for s in meta["key_sheets"] if s in available]

    if not target_sheets:
        target_sheets = available[:4]

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        md_table = _sheet_to_markdown(ws, max_rows=meta.get("max_rows", 100))
        if md_table.strip():
            sections.append(f"## {sheet_name}\n\n{md_table}\n")

    return "\n".join(sections)


# ─────────────────────────────────────────────────────────────────────────────
# Discovery and batch extraction
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_SEARCH_PATHS = [
    Path.home() / "Desktop",
    Path.home() / "Downloads",
    Path.home() / "Documents" / "GitHub",
    Path.home() / "Documents",
]


def discover_and_extract(output_dir: Path, search_paths: Optional[List[Path]] = None) -> None:
    if search_paths is None:
        search_paths = DEFAULT_SEARCH_PATHS

    output_dir.mkdir(parents=True, exist_ok=True)
    found = 0
    skipped = 0

    for entry in WORKBOOK_REGISTRY:
        path = _find_workbook(entry["pattern"], search_paths)
        if path is None:
            print(f"  [skip] {entry['pattern']} — not found in search paths")
            skipped += 1
            continue

        out_path = output_dir / f"{entry['slug']}.md"
        print(f"  [extract] {path.name} → {out_path.name}")
        try:
            content = extract_workbook(path, entry)
            out_path.write_text(content, encoding="utf-8")
            found += 1
        except Exception as e:
            print(f"    ERROR: {e}")
            skipped += 1

    print(f"\nDone: {found} extracted, {skipped} skipped → {output_dir}")


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract xlsx workbooks to structured Markdown knowledge files"
    )
    parser.add_argument(
        "--output-dir",
        default=str(
            Path(__file__).parent.parent / "knowledge-extracts"
        ),
        help="Directory to write extracted .md files (default: ../knowledge-extracts/)",
    )
    parser.add_argument(
        "--search-path",
        action="append",
        dest="search_paths",
        help="Additional directory to search for xlsx files (repeatable)",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    search_paths = list(DEFAULT_SEARCH_PATHS)
    if args.search_paths:
        search_paths = [Path(p) for p in args.search_paths] + search_paths

    print(f"Output directory: {output_dir}")
    print(f"Search paths: {[str(p) for p in search_paths]}\n")

    discover_and_extract(output_dir, search_paths)


if __name__ == "__main__":
    main()
